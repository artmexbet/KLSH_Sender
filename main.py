import smtplib
import sys
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from PyQt5.QtWidgets import QMainWindow, QApplication, QDialog
from PyQt5.QtWidgets import QFileDialog
from openpyxl.worksheet.worksheet import Worksheet

from design.MainWindow import Ui_MainWindow
from design.ColumnDialog import Ui_Dialog
from design.help import Ui_Help

from openpyxl import load_workbook, Workbook

import json
import os

DIRECTIONS = {"НТН": "направление точных наук",
              "НЕН": "направление естественных",
              "НОН": "направление общественных наук",
              "НФН": "направление филологических наук"}


class Student:
    def __init__(self, name, direction, place, email):
        self.name = name
        self.directions = [DIRECTIONS[direction]]
        self._places = [place]
        self.email = email

    def add_direction(self, new_dir):
        self.directions.append(DIRECTIONS[new_dir])

    def add_place(self, place):
        self._places.append(place)

    @property
    def direction(self):
        return ", ".join(self.directions).capitalize()

    @property
    def places(self):
        return ", ".join([str(i) for i in self._places])


class Config:
    def __init__(self):
        if not os.path.exists("config.json"):
            with open("config.json", "w", encoding="utf8") as f:
                json.dump({"login": "", "name_col": -1, "dir_col": -1,
                           "place_col": -1, "email_col": -1},
                          f, ensure_ascii=False)
        with open("config.json", encoding="utf8") as f:
            self.cfg_dict = json.load(f)

    @property
    def login(self) -> str:
        return self.cfg_dict["login"] if self.cfg_dict["login"] else None

    @property
    def name_col(self) -> int:
        return self.cfg_dict["name_col"] if self.cfg_dict[
                                                "name_col"] != -1 else None

    @property
    def dir_col(self) -> int:
        return self.cfg_dict["dir_col"] if self.cfg_dict[
                                               "dir_col"] != -1 else None

    @property
    def place_col(self) -> int:
        return self.cfg_dict["place_col"] if self.cfg_dict[
                                                 "place_col"] != -1 else None

    @property
    def email_col(self) -> int:
        return self.cfg_dict["email_col"] if self.cfg_dict[
                                                 "email_col"] != -1 else None

    @login.setter
    def login(self, value: str):
        self.cfg_dict["login"] = value
        self.commit()

    @name_col.setter
    def name_col(self, value: int):
        self.cfg_dict["name_col"] = value - 1
        self.commit()

    @dir_col.setter
    def dir_col(self, value: int):
        self.cfg_dict["dir_col"] = value - 1
        self.commit()

    @place_col.setter
    def place_col(self, value: int):
        self.cfg_dict["place_col"] = value - 1
        self.commit()

    @email_col.setter
    def email_col(self, value: int):
        self.cfg_dict["email_col"] = value - 1
        self.commit()

    def commit(self):
        with open("config.json", "w", encoding="utf8") as f:
            json.dump(self.cfg_dict, f, ensure_ascii=False)


class HelpDialog(QDialog, Ui_Help):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("Справка")


class ColumnDialog(QDialog, Ui_Dialog):
    def __init__(self, config: Config):
        super().__init__()
        self.setupUi(self)

        self.is_ok = False

        self.name_column.setText(
            str(config.name_col + 1 if config.name_col is not None else ""))
        self.dir_column.setText(
            str(config.dir_col + 1 if config.dir_col is not None else ""))
        self.email_column.setText(
            str(config.email_col + 1 if config.email_col is not None else ""))
        self.place_column.setText(
            str(config.place_col + 1 if config.place_col is not None else ""))

        self.ok_btn.clicked.connect(self.ok_action)
        self.cancel_btn.clicked.connect(self.close)

    def ok_action(self):
        self.is_ok = True
        self.close()

    @property
    def name_col(self) -> int:
        return int(self.name_column.text())

    @property
    def direction_col(self) -> int:
        return int(self.dir_column.text())

    @property
    def place_col(self) -> int:
        return int(self.place_column.text())

    @property
    def table_name(self) -> str:
        return self.sheet_name.text().strip()

    @property
    def email_col(self) -> int:
        return int(self.email_column.text())


class MainWindow(QMainWindow, Ui_MainWindow):
    excel_table: Worksheet

    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.choose_file_btn.clicked.connect(self.get_xl_file)
        self.choose_html_file_btn.clicked.connect(self.get_html_file)
        self.choose_txt_file_btn.clicked.connect(self.get_txt_file)
        self.send.clicked.connect(self.send_messages)

        self.excel_table = None
        self.message_text = None
        self.alt_message_text = None

        self.help_dialog = None

        self.config = Config()

        self.login_input.setText(self.config.login)

        self.open_about.triggered.connect(self.open_help)

        self.setWindowTitle("Рассылка")

    def open_help(self):
        self.help_dialog = HelpDialog()
        self.help_dialog.show()

    def get_xl_file(self):
        filename = QFileDialog.getOpenFileName(self, "Выберите файл", "",
                                               "Excel files (*.xlsx *.xls)")[0]
        if filename:
            dialog = ColumnDialog(self.config)
            dialog.exec()
            if dialog.is_ok:
                wb = load_workbook(filename)
                self.excel_table = wb[dialog.table_name]
                self.config.name_col = dialog.name_col
                self.config.dir_col = dialog.direction_col
                self.config.place_col = dialog.place_col  # TODO: Разобраться, в каком виде хранятся места в таблице
                self.config.email_col = dialog.email_col
                self.excel_filename_label.setText(filename)

    def get_html_file(self):
        filename = QFileDialog.getOpenFileName(self, "Выберите файл", "",
                                               "HTML files (*.html)")[0]
        with open(filename, encoding="utf8") as f:
            self.message_text = f.read()
            self.html_filename_label.setText(filename)

    def get_txt_file(self):
        filename = QFileDialog.getOpenFileName(self, "Выберите файл", "",
                                               "Text files (*.txt)")[0]
        with open(filename, encoding="utf8") as f:
            self.alt_message_text = f.read()
            self.txt_filename_label.setText(filename)

    def send_messages(self):
        if not self.excel_table:
            self.statusbar.showMessage("Не прикреплена таблица Excel")
        elif not self.message_text:
            self.statusbar.showMessage(
                "Не прикреплён HTML файл или прикреплён пустой")
        elif not self.login_input.text() or not self.password_input.text():
            self.statusbar.showMessage("Не вписан логин или пароль")

        self.config.login = self.login_input.text()

        try:
            smtp = smtplib.SMTP("smtp.gmail.com", 587)
            smtp.starttls()
            smtp.login(self.config.login, self.password_input.text())
        except Exception as ex:
            print(ex)  # Если авторизация не проходит, то мы крашим функцию
            self.statusbar.showMessage("Неверный логин или пароль")
            return

        # Собираем инфу из таблицы по каждому школьнику
        # (у некоторых школьников есть достижения не по одному направлению)
        students = {}
        for row in self.excel_table.rows:
            name = row[self.config.name_col].value
            # TODO: Убедиться, что у таблицы школьников именно такая структура
            #  (чтобы грамотно подставлять в письмо строки)
            if name not in students.keys():
                students[name] = Student(name,
                                         row[self.config.dir_col].value,
                                         row[self.config.place_col].value,
                                         row[self.config.email_col].value)
            else:
                students[name].add_direction(row[self.config.dir_col].value)
                students[name].add_place(row[self.config.place_col].value)

        wb = Workbook()  # Создаём таблицу для тех школьников, которым не удалось отправить письма
        ws = wb.create_sheet("Invalid Students", 0)
        ws.cell(1, 1, "Имя")
        ws.cell(1, 2, "Направления")
        ws.cell(1, 3, "Места")
        ws.cell(1, 4, "Почта")

        for name, student in students.items():
            temp_text = self.message_text.replace("{{ name }}", name).replace(
                "{{ dir }}", student.direction).replace("{{ place }}",
                                                        student.places)
            # TODO: Разобраться с тем, как вписаны в таблицу места, чтобы правильно подставлять места, и с тем,
            #   как форматировать имя (возможно, нужно сплитать)

            message = MIMEMultipart("alternative")
            message["Subject"] = self.message_subject.text()
            message["From"] = "Красноярская летняя школа"
            message["To"] = student.email
            part = MIMEText(temp_text, "html")
            message.attach(MIMEText(self.alt_message_text, "plain"))
            message.attach(part)

            try:
                smtp.sendmail(self.config.login, student.email,
                              message.as_string())
            except Exception as ex:
                # В случае, если сообщение не было отправлено,
                # школьника добавит в файл
                print(ex)
                ws.cell(ws.max_row + 1, 1, student.name)
                ws.cell(ws.max_row + 1, 2, student.direction)
                ws.cell(ws.max_row + 1, 3, student.places)
                ws.cell(ws.max_row + 1, 4, student.email)
        wb.save("troubles.xlsx")


def exception_hook(exctype, value, traceback):
    print(exctype, value, traceback)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    sys.excepthook = exception_hook
    window = MainWindow()
    window.show()
    app.exec()
