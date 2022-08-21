import openpyxl

wb = openpyxl.load_workbook("test.xlsx")
sheet = wb.active

print(sheet.max_row)
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(type(sheet.cell(i, j).value), end=" ")
    print()
